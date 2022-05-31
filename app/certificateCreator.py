from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import os
import zipfile


def create_single_sertificate(file, hight=1250, text_size=150, name='Ivanov Ivan Ivanovich', name2=""):
    #  шрифт
    # font = ImageFont.truetype(font_file_path, text_size)
    font = ImageFont.truetype("arial.ttf", text_size)
    fontNum = ImageFont.truetype("arial.ttf", 30)

    # загружаете фоновое изображение
    img = Image.open(file)
    w, h = img.size

    img2 = img.copy()
    # определяете объект для рисования
    draw = ImageDraw.Draw(img2)
    # цвет текста, RGB
    text_color = (111, 155, 182)
    num = "№341/2022"
    # добавляем текст
    # draw.text((w/2, hight), name, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
    # draw.text((w / 2, hight + 100), name2, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
    draw.text((1400, hight), name, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
    draw.text((1400, hight + 75), name2, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
    draw.text((350, 1400), num, (0, 0, 0), fontNum, anchor="ms", stroke_width=1, stroke_fill=text_color)

    # сохраняем новое изображение
    if name2:
        img2.save(f"app\\uploads\\{name} {name2}.png")
        return f"uploads\\{name} {name2}.png"
    else:
        img2.save(f"app\\uploads\\{name}.png")
        return f"uploads\\{name}.png"



def create_list_sertificate(file, excel_file, path_name='heap', hight=1250, text_size=150, name='Ivanov Ivan Ivanovich', name2=""):
    #  шрифт
    # font = ImageFont.truetype(font_file_path, text_size)
    font = ImageFont.truetype("arial.ttf", text_size)
    fontNum = ImageFont.truetype("arial.ttf", 30)

    # загружаете фоновое изображение
    img = Image.open(file)
    w, h = img.size
    # цвет текста, RGB
    # text_color = (111, 155, 182)
    text_color = (0, 153, 159)

    wb_obj = openpyxl.load_workbook(excel_file)
    sheet_obj = wb_obj.active
    for i in range(1, sheet_obj.max_row + 1):
        surname = sheet_obj.cell(row=i, column=1).value
        name_ = sheet_obj.cell(row=i, column=2).value
        name = f"{surname} {name_}"
        name2 = sheet_obj.cell(row=i, column=3).value
        num = f"№{sheet_obj.cell(row=i, column=5).value}"
        print(surname)

        img2 = img.copy()

        # определяете объект для рисования
        draw = ImageDraw.Draw(img2)

        # добавляем текст
        # draw.text((w/2, hight), name, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
        # draw.text((w / 2, hight + 100), name2, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
        draw.text((1400, hight), name, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
        draw.text((1400, hight + 75), name2, text_color, font, anchor="ms", stroke_width=1, stroke_fill=text_color)
        draw.text((350, 1400), num, (0, 0, 0), fontNum, anchor="ms", stroke_width=1, stroke_fill=text_color)

        # сохраняем новое изображение

        if not os.path.exists(f"app\\uploads\\{path_name}"):
            os.makedirs(f"app\\uploads\\{path_name}")
        img2.save(f"app\\uploads\\{path_name}\\{name} {name2}.png")

        # dtn = datetime.now().strftime("%d-%m-%Y-%H-%M")
        # if not os.path.exists(f"app\\uploads\\{dtn}"):
        #     os.makedirs(f"app\\uploads\\{dtn}")
        # if name2:
        #     img2.save(f"app\\uploads\\{dtn}\\{name} {name2}.png")
        # else:
        #     img2.save(f"app\\uploads\\{dtn}\\{name}.png")

    # with zipfile.ZipFile(f"app\\uploads\\{dtn}.zip", mode='w') as zipf:
    #     len_dir_path = len(f"app\\uploads\\{dtn}")
    #     for root, _, files in os.walk(f"app\\uploads\\{dtn}"):
    #         for file in files:
    #             file_path = os.path.join(root, file)
    #             zipf.write(file_path, file_path[len_dir_path:])
    #             wb_obj.close()
    # return f"uploads\\{dtn}.zip"


